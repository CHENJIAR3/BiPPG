from __future__ import annotations
import os
import datetime
import itertools
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from scipy import stats as scipy_stats
from scikit_posthocs import posthoc_dunn
from presentions.plot_setting import set_plot_style

set_plot_style()

# ══════════════════════════════════════════════════════════════════
#  配置
# ══════════════════════════════════════════════════════════════════
formatted_date = datetime.datetime.now().strftime("%Y-%m-%d")

GROUPS = {
    "age":      [[0, 60],[60, 75], [75, 1000]],
    "bmi":      [[0 , 23.9], [23.9, 27.9], [27.9, 100]],
    "gender":   [[1], [2]],
    "position": [[1], [2], [3]],
}
GROUP_LABELS = {
    "age":      ["<60", "60–75", "≥75"],
    "bmi":      ["<23.9", "23.9–27.9", "≥27.9"],
    "gender":   ["Male", "Female"],
    "position": ["Lay", "Sit", "Stand"],
}
TARGET_COLS = ["sbp_fix", "dbp_fix", "pr_ref", "age"]
PRED_L_COLS = ["pred_sbp", "pred_dbp", "pred_hr", "pred_age"]
PRED_R_COLS = ["pred_sbp_right", "pred_dbp_right",
               "pred_hr_right", "pred_age_right"]
METRICS     = ["SBP_MAE", "DBP_MAE", "HR_MAE", "Age_MAE"]
UNITS       = ["mmHg", "mmHg", "bpm", "year"]
_CONFIGS    = ["BD"]

# ── 配色：每个亚组用渐变色系，衬托组间差异 ─────────────────────────
_GROUP_PALETTES = {
    "age":      ["#C6DBEF", "#9ECAE1", "#6BAED6", "#3182BD", "#08519C"],
    "bmi":      ["#C7E9C0", "#A1D99B", "#74C476", "#31A354"],
    "gender":   ["#FCBBA1", "#CB181D"],
    "position": ["#DADAEB", "#9E9AC8", "#6A51A3"],
}


# ══════════════════════════════════════════════════════════════════
#  统计核心（不变）
# ══════════════════════════════════════════════════════════════════
def _mae(a, b): return np.mean(np.abs(a - b), axis=0)
def _std(a, b): return np.std(np.abs(a - b), axis=0, ddof=1)


def compute_subgroup_stats(result_df: pd.DataFrame) -> dict:
    stats = {}

    def _add(key, sub):
        if len(sub) == 0:
            return
        pl = sub[PRED_L_COLS].values
        pr = sub[PRED_R_COLS].values
        tg = sub[TARGET_COLS].values
        stats[key] = {
            "n":      len(sub),
            "mae_BD": _mae(pl, pr),
            "std_BD": _std(pl, pr),
            "err_BD": np.abs(pl - pr),   # (N, 4)
        }

    _add("All", result_df)
    for gk, gvals in GROUPS.items():
        for gv, label in zip(gvals, GROUP_LABELS[gk]):
            if len(gv) == 2:
                sub = result_df[
                    (result_df[gk] >= gv[0]) & (result_df[gk] < gv[1])]
            else:
                sub = result_df[result_df[gk] == gv[0]]
            _add(f"{gk}_{label}", sub)
    return stats


# ══════════════════════════════════════════════════════════════════
#  统计检验
# ══════════════════════════════════════════════════════════════════
def run_kruskal_dunn(groups_data: list) -> tuple:
    groups_data = [g for g in groups_data if len(g) > 0]
    if len(groups_data) < 2:
        return np.nan, None
    kw_stat, kw_p = scipy_stats.kruskal(*groups_data)
    if len(groups_data) == 2:
        _, mw_p = scipy_stats.mannwhitneyu(
            groups_data[0], groups_data[1], alternative="two-sided")
        dunn = pd.DataFrame([[1.0, mw_p], [mw_p, 1.0]])
        return kw_p, dunn
    all_vals = np.concatenate(groups_data)
    all_grp  = np.concatenate([
        np.full(len(g), i) for i, g in enumerate(groups_data)])
    dunn = posthoc_dunn(
        pd.DataFrame({"val": all_vals, "grp": all_grp}),
        val_col="val", group_col="grp", p_adjust="bonferroni")
    return kw_p, dunn


# ══════════════════════════════════════════════════════════════════
#  统计检验结果收集（新增全局容器）
# ══════════════════════════════════════════════════════════════════
_STAT_RECORDS: list[dict] = []   # 运行结束后统一保存


def p_to_stars(p: float) -> str:
    if np.isnan(p) or p >= 0.05: return "ns"   # ← 改为返回 "ns"
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    return "*"


# ══════════════════════════════════════════════════════════════════
#  显著性标注（层级 bracket）
# ══════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════
#  显著性标注（支持 ns 展示，样式分开）
# ══════════════════════════════════════════════════════════════════
def _annotate_significance(ax, x1, x2, y_top, stars,
                            color="#333333", linewidth=0.7, fontsize=12):
    if not stars :
        return

    yrange = ax.get_ylim()[1] - ax.get_ylim()[0]
    h      = yrange * 0.025

    # ns 用浅灰虚线，显著用实线
    ls     = "--" if stars == "ns" else "-"
    fc     = color
        # if stars == "ns" else color
    ax.plot([x1, x1, x2, x2],
            [y_top, y_top + h, y_top + h, y_top],
            lw=linewidth, color=fc, linestyle=ls, clip_on=False)
    ax.text((1.25*x1 + 0.75*x2) / 2, y_top + h , stars,
            ha="center", va="bottom",
            fontsize=fontsize if stars != "ns" else fontsize - 0.5,
            color=fc,
            # style="italic" if stars == "ns" else "normal",
            clip_on=False)

# ══════════════════════════════════════════════════════════════════
#  Ax 样式
# ══════════════════════════════════════════════════════════════════
def _setup_ax(ax, ylabel, xtick_labels, rotate=30):
    ax.set_ylabel(ylabel)
    ax.set_xticks(range(len(xtick_labels)))
    ax.set_xticklabels(xtick_labels, rotation=rotate, ha="center")
    ax.tick_params(axis="x", length=0)
    ax.yaxis.grid(True, linewidth=0.4, linestyle=":",
                  color="gray", alpha=0.5, zorder=0)
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(0.6)
    ax.spines["bottom"].set_linewidth(0.6)
    # x 轴下方留空间给 n= 标注
    # ax.tick_params(axis="x", pad=14)

# ══════════════════════════════════════════════════════════════════
#  核心绘图
# ══════════════════════════════════════════════════════════════════
def plot_group_bar_with_stats(stats: dict,
                               # result_df: pd.DataFrame,
                               group_key: str,
                               model_name: str,
                               save_dir: str):
    stat_keys  = [f"{group_key}_{lb}" for lb in GROUP_LABELS[group_key]
                  if f"{group_key}_{lb}" in stats]
    sub_labels = [k.split("_", 1)[1] for k in stat_keys]
    n_grp      = len(stat_keys)
    x          = np.arange(n_grp)
    bar_colors = _GROUP_PALETTES[group_key][:n_grp]   # 每亚组独立颜色

    for metric, unit in zip(METRICS, UNITS):
        mi           = METRICS.index(metric)
        metric_short = metric.replace("_MAE", "")

        fig, ax = plt.subplots(figsize=(6, 5), constrained_layout=True)

        means = [stats[k]["mae_BD"][mi] for k in stat_keys]
        stds  = [stats[k]["std_BD"][mi] for k in stat_keys]

        # ── 柱状图：每亚组独立色 ────────────────────────────────
        bars = ax.bar(x, means,
                      width=0.55,
                      color=bar_colors,
                      edgecolor="white",
                      linewidth=2.5,
                      alpha=0.92,
                      yerr=stds,
                      error_kw=dict(elinewidth=0.8, capsize=3,
                                    capthick=0.8, ecolor="#444",
                                    alpha=0.6),
                      zorder=3)

        # # ── 柱顶数值标注 ─────────────────────────────────────────
        # for xi, m, s in zip(x, means, stds):
        #     ax.text(xi, m + s - (max(means) - min(means)) * 0.2,
        #             f"{m:.2f}",
        #             ha="center", va="bottom",
        #              # color="#333333"
        #             )

        # ── 样本量标注（x轴下方）────────────────────────────────
        for xi, k in zip(x, stat_keys):
            n = stats[k]["n"]
            ax.text(xi, ax.get_ylim()[0] - (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.1,
                    f"n={n:,}",
                    ha="center", va="top",
                    # fontsize=5.5,
                    # color="#666666",
                    clip_on=False)

        # ── Kruskal-Wallis + Dunn 标注 ──────────────────────────
        groups_err = [stats[k]["err_BD"][:, mi] for k in stat_keys]
        kw_p, dunn = run_kruskal_dunn(groups_err)
        stars_kw   = p_to_stars(kw_p)

        if dunn is not None and stars_kw:
            bar_tops = [m + s for m, s in zip(means, stds)]
            yrange   = ax.get_ylim()[1] - ax.get_ylim()[0]
            level_h  = yrange * 0.09
            occupied = {}
            #
            # pairs = list(itertools.combinations(range(n_grp), 2))
            # pairs.sort(key=lambda ij: dunn.iloc[ij[0], ij[1]])
            pairs = list(itertools.combinations(range(n_grp), 2))
            # 显著的排前面（层级低），ns 排后面
            # pairs.sort(key=lambda ij: dunn.iloc[ij[0], ij[1]])

            for i, j in pairs:
                p_ij  = dunn.iloc[i, j]
                s_ij  = p_to_stars(p_ij)

                level = max((occupied.get(k, 0) for k in range(i, j + 1)),
                            default=0)
                y_ann = max(bar_tops[i], bar_tops[j]) + \
                        level * level_h + yrange * 0.04

                _annotate_significance(ax, x[i], x[j], y_ann, s_ij,
                                       color="#555555")
                for k in range(i, j + 1):
                    occupied[k] = level + 1

                # ── 收集统计结果 ──────────────────────────────
                _STAT_RECORDS.append({
                    "model":      model_name,
                    "group_key":  group_key,
                    "metric":     metric_short,
                    "subgroup_A": sub_labels[i],
                    "subgroup_B": sub_labels[j],
                    "n_A":        stats[stat_keys[i]]["n"],
                    "n_B":        stats[stat_keys[j]]["n"],
                    "KW_p":       float(kw_p),
                    "KW_stars":   stars_kw,
                    "Dunn_p":     float(p_ij),
                    "Dunn_stars": s_ij,
                    "KW_p_display": format_p(kw_p),

                    "Dunn_p_display": format_p(p_ij),
                })
            # 动态扩展 y 上限
            if occupied:
                extra = max(occupied.values()) * level_h * 0.6
                ax.set_ylim(ax.get_ylim()[0],
                            ax.get_ylim()[1] + extra)

        # # ── KW 整体 p 值标注（右上角）──────────────────────────
        # if not np.isnan(kw_p):
        #     p_txt = (f"KW p {'< 0.001' if kw_p < 0.001 else f'= {kw_p:.3f}'}")
        #     ax.text(0.98, 0.98, p_txt,
        #             transform=ax.transAxes,
        #             ha="right", va="top",
        #             fontsize=5.5, color="#666666",
        #             style="italic")
        _setup_ax(ax, f"BD {metric_short} ({unit})", sub_labels, rotate=0)

        # _setup_ax(ax, f"{metric_short}({unit})", sub_labels, rotate=0)
        #
        # # 星号说明
        # ax.text(0.01, 0.99,
        #         "*p<0.05  **p<0.01  ***p<0.001",
        #         transform=ax.transAxes,
        #         va="top", ha="left",
        #         fontsize=5, color="#888888")
        # # 星号图例
        # ax.text(0.01, 0.99,
        #         "ns p≥0.05  * p<0.05  ** p<0.01  *** p<0.001",
        #         transform=ax.transAxes,
        #         va="top", ha="left", fontsize=5, color="#888888")
        fname = os.path.join(
            save_dir, f"BD_{group_key}_{metric_short}.pdf")
        fname2 = os.path.join(
            save_dir, f"BD_{group_key}_{metric_short}.png")
        print(fname)
        fig.savefig(fname, dpi=600, bbox_inches="tight")
        fig.savefig(fname2, dpi=600, bbox_inches="tight")
        # plt.show()
        plt.close(fig)
        print(f"  📊 {fname}")


# ══════════════════════════════════════════════════════════════════
#  Excel
# ══════════════════════════════════════════════════════════════════
def _apply_excel_style(ws, df: pd.DataFrame):
    HDR_FILL  = PatternFill("solid", fgColor="1F497D")
    GRP_FILLS = {
        "age":      PatternFill("solid", fgColor="DEEAF1"),
        "bmi":      PatternFill("solid", fgColor="E2EFDA"),
        "gender":   PatternFill("solid", fgColor="FFF2CC"),
        "position": PatternFill("solid", fgColor="FCE4D6"),
        "All":      PatternFill("solid", fgColor="F2F2F2"),
    }
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=8)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border

    group_col_idx = df.columns.get_loc("Group") + 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        gk   = ws.cell(row=row_idx, column=group_col_idx).value or "All"
        fill = GRP_FILLS.get(gk, GRP_FILLS["All"])
        for cell in row:
            cell.fill      = fill
            cell.border    = border
            cell.font      = Font(name="Arial", size=8)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = {"Model": 14, "Group": 10, "Subgroup": 12,
                  "Config": 14, "N": 7}
    for i, col in enumerate(df.columns, 1):
        w = col_widths.get(col, 14 if "display" in col else 9)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def stats_to_dataframe(stats: dict, model_name: str) -> pd.DataFrame:
    rows = []
    for key, s in stats.items():
        group_key, *rest = key.split("_", 1)
        subgroup = rest[0] if rest else "All"
        row = {"Model": model_name, "Group": group_key,
               "Subgroup": subgroup, "Config": "BD", "N": s["n"]}
        for i, m in enumerate(METRICS):
            row[m]              = round(s["mae_BD"][i], 3)
            row[m + "_std"]     = round(s["std_BD"][i], 3)
            row[m + "_display"] = f"{s['mae_BD'][i]:.2f}±{s['std_BD'][i]:.2f}"
        rows.append(row)
    return pd.DataFrame(rows)

def format_p(p):
    if np.isnan(p):
        return "NA"
    if p < 1e-3:
        return f"{p:.2e}"   # 科学计数法
    else:
        return f"{p:.4f}"
# ══════════════════════════════════════════════════════════════════
#  主入口
# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    pkl_dir   = "/home/cjr/PPGBen/predictions/"
    save_dir  = "../figures/subgroup_BD/"
    save_xlsx = f"../results/subpopulation_results_{formatted_date}.xlsx"
    os.makedirs(save_dir, exist_ok=True)

    order = ["CRNN", "ACNN", "ResNet1D", "Net1D", "LSTM", "Efficient1D",
             "AutoFormer", "PatchTST", "Informer", "iTransformer",
             "ResNet1DMoE", "CSFM", "PaAno"]
    pkl_files = sorted(os.listdir(pkl_dir),
                       key=lambda x: order.index(x.split("_")[0]))

    all_dfs = {}
    for pkl_file in pkl_files[3:4]:
        model_name = pkl_file.split("_")[0]
        result_all = pd.read_pickle(os.path.join(pkl_dir, pkl_file))
        print(f"\n▶ {model_name}  (n={len(result_all)})")

        stats = compute_subgroup_stats(result_all)
        df    = stats_to_dataframe(stats, model_name)
        all_dfs[model_name] = df

        for gk in GROUPS.keys():
            plot_group_bar_with_stats(
                stats, gk, model_name, save_dir)
    df = pd.DataFrame(all_dfs["Net1D"])
    df_stat = pd.DataFrame(_STAT_RECORDS)
    # 按 model / group / metric / Dunn_p 排序
    df_stat = df_stat.sort_values(
        ["model", "group_key", "metric", "Dunn_p"]).reset_index(drop=True)
    save_path = os.path.join(f"../results/subpopulation_BD_{formatted_date}.xlsx")
    # writer = pd.ExcelWriter(save_path, engine="xlsxwriter")
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:

        df_stat.to_excel(writer, sheet_name="stat", index=False)

        df.to_excel(writer, sheet_name="result", index=False)