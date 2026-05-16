"""
Author: jiarong chen
2026/05/15
# subpopulation analyasis BD in subject level
"""
from __future__ import annotations
import os
import datetime
import itertools
import glob
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy import stats as scipy_stats
from scikit_posthocs import posthoc_dunn

# ── 如有自定义绘图风格，取消注释 ──────────────────────────────────────────────
# from presentions.plot_setting import set_plot_style
# set_plot_style()

formatted_date = datetime.datetime.now().strftime("%Y-%m-%d")

# ══════════════════════════════════════════════════════════════════
#  配置
# ══════════════════════════════════════════════════════════════════
GROUPS = {
    "age":      [[0, 60], [60, 75], [75, 1000]],
    "bmi":      [[0, 23.9], [23.9, 27.9], [27.9, 100]],
    "gender":   [[1], [2]],
    "position": [[1], [2], [3]],
}
GROUP_LABELS = {
    "age":      ["<60", "60–75", "≥75"],
    "bmi":      ["<23.9", "23.9–27.9", "≥27.9"],
    "gender":   ["Male", "Female"],
    "position": ["Lay", "Sit", "Stand"],
}

TARGET_COLS = ["sbp_fix", "dbp_fix", "pr_ref", "age"]
PRED_L_COLS = ["pred_sbp",       "pred_dbp",       "pred_hr",       "pred_age"]
PRED_R_COLS = ["pred_sbp_right", "pred_dbp_right",  "pred_hr_right", "pred_age_right"]
METRICS     = ["SBP", "DBP", "HR", "Age"]
UNITS       = ["mmHg", "mmHg", "bpm", "year"]

_GROUP_PALETTES = {
    "age":      ["#C6DBEF", "#6BAED6", "#08519C"],
    "bmi":      ["#C7E9C0", "#74C476", "#238B45"],
    "gender":   ["#FCBBA1", "#CB181D"],
    "position": ["#DADAEB", "#9E9AC8", "#6A51A3"],
}

MODELNAMES = [
    "CRNN", "ACNN", "ResNet1D", "Net1D", "LSTM", "Efficient1D",
    "AutoFormer", "PatchTST", "Informer", "iTransformer",
    "ResNet1DMoE", "CSFM",
]


# ══════════════════════════════════════════════════════════════════
#  变量类型声明
#  SUBJECT_VARS：受试者固定属性，subject level 切片
#  SEGMENT_VARS：每段数据属性，必须在 segment level 切片后再聚合
# ══════════════════════════════════════════════════════════════════
SUBJECT_VARS = ["age", "bmi", "gender"]   # 一人一个值
SEGMENT_VARS = ["position"]               # 一人可有多个值，不能用 first 聚合

BD_ERR_COLS = [f"bd_{m.lower()}" for m in METRICS]


def _add_bd_errors(df: pd.DataFrame) -> pd.DataFrame:
    """在 segment level 计算 |pred_L − pred_R|，原地添加列，返回副本。"""
    df = df.copy()
    for pred_l, pred_r, bd_col in zip(PRED_L_COLS, PRED_R_COLS, BD_ERR_COLS):
        df[bd_col] = (df[pred_l] - df[pred_r]).abs()
    return df


def make_subject_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    用于 subject-level 变量（age/bmi/gender）的聚合。
    先算 segment-level |pred_L − pred_R|，再 groupby subject 取均值。
    position 不在此处聚合（它是 segment-level 属性）。
    """
    df = _add_bd_errors(df)

    agg_dict = {col: "mean" for col in BD_ERR_COLS}
    for gv in SUBJECT_VARS:
        if gv in df.columns:
            agg_dict[gv] = "first"   # 固定属性，取 first 安全

    return df.groupby("subject_id").agg(agg_dict).reset_index()


def make_segment_subgroup_df(df: pd.DataFrame, seg_var: str,
                              gv: list, label: str) -> pd.DataFrame:
    """
    用于 segment-level 变量（position）的亚组聚合。
    先按 seg_var 值过滤 segments，再在该子集里 groupby subject 取均值。
    这样 n 是"在该姿势下有数据的受试者数"，且同一人可出现在多个亚组里。
    """
    df = _add_bd_errors(df)

    if len(gv) == 2:
        sub_seg = df[(df[seg_var] >= gv[0]) & (df[seg_var] < gv[1])]
    else:
        sub_seg = df[df[seg_var] == gv[0]]

    if len(sub_seg) == 0:
        return pd.DataFrame()

    return sub_seg.groupby("subject_id")[BD_ERR_COLS].mean().reset_index()


# ══════════════════════════════════════════════════════════════════
#  Step 2：亚组统计量收集
# ══════════════════════════════════════════════════════════════════
def compute_subgroup_stats(raw_df: pd.DataFrame) -> dict:
    """
    raw_df：原始 segment-level DataFrame。
    对 subject-level 变量在 subject_df 上切片；
    对 segment-level 变量（position）先过滤 segments 再聚合。
    """
    stats = {}
    subject_df = make_subject_df(raw_df)   # 供 subject-level 变量使用

    def _add(key: str, sub: pd.DataFrame):
        if len(sub) == 0:
            return
        err = sub[BD_ERR_COLS].values
        stats[key] = {
            "n":   len(sub),
            "mae": err.mean(axis=0),
            "std": err.std(axis=0, ddof=1),
            "err": err,
        }

    _add("All", subject_df)

    for gk, gvals in GROUPS.items():
        for gv, label in zip(gvals, GROUP_LABELS[gk]):
            key = f"{gk}_{label}"

            if gk in SUBJECT_VARS:
                # ── subject-level 切片 ────────────────────────
                if gk not in subject_df.columns:
                    continue
                if len(gv) == 2:
                    sub = subject_df[
                        (subject_df[gk] >= gv[0]) & (subject_df[gk] < gv[1])]
                else:
                    sub = subject_df[subject_df[gk] == gv[0]]
                _add(key, sub)

            else:
                # ── segment-level 切片（position 等）────────────
                sub = make_segment_subgroup_df(raw_df, gk, gv, label)
                _add(key, sub)

    return stats


# ══════════════════════════════════════════════════════════════════
#  Step 3：统计检验（Kruskal-Wallis + Dunn/Mann-Whitney）
# ══════════════════════════════════════════════════════════════════
def run_kruskal_dunn(groups_data: list[np.ndarray]) -> tuple:
    groups_data = [g for g in groups_data if len(g) > 0]
    if len(groups_data) < 2:
        return np.nan, None
    kw_stat, kw_p = scipy_stats.kruskal(*groups_data)
    if len(groups_data) == 2:
        _, mw_p = scipy_stats.mannwhitneyu(
            groups_data[0], groups_data[1], alternative="two-sided")
        dunn = pd.DataFrame([[1.0, mw_p], [mw_p, 1.0]])
        return kw_p, dunn
    all_vals = np.concatenate(groups_data)
    all_grp  = np.concatenate([
        np.full(len(g), i) for i, g in enumerate(groups_data)])
    dunn = posthoc_dunn(
        pd.DataFrame({"val": all_vals, "grp": all_grp}),
        val_col="val", group_col="grp", p_adjust="bonferroni")
    return kw_p, dunn


def p_to_stars(p: float) -> str:
    if np.isnan(p) or p >= 0.05: return "ns"
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    return "*"


def format_p(p: float) -> str:
    if np.isnan(p): return "NA"
    if p < 1e-3:    return f"{p:.2e}"
    return f"{p:.4f}"


# ══════════════════════════════════════════════════════════════════
#  Step 4：显著性标注
# ══════════════════════════════════════════════════════════════════
def _annotate_significance(ax, x1, x2, y_top, stars,
                            color="#333333", linewidth=0.7, fontsize=7):
    if not stars:
        return
    yrange = ax.get_ylim()[1] - ax.get_ylim()[0]
    h      = yrange * 0.025
    ls     = "--" if stars == "ns" else "-"
    ax.plot([x1, x1, x2, x2],
            [y_top, y_top + h, y_top + h, y_top],
            lw=linewidth, color=color, linestyle=ls, clip_on=False)
    ax.text((x1 + x2) / 2, y_top + h, stars,
            ha="center", va="bottom",
            fontsize=fontsize if stars != "ns" else fontsize - 0.5,
            color=color, clip_on=False)


def _setup_ax(ax, ylabel, xtick_labels, rotate=0):
    ax.set_ylabel(ylabel)
    ax.set_xticks(range(len(xtick_labels)))
    ax.set_xticklabels(xtick_labels, rotation=rotate, ha="center")
    ax.tick_params(axis="x", length=0)
    ax.yaxis.grid(True, linewidth=0.4, linestyle=":", color="gray",
                  alpha=0.5, zorder=0)
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(0.6)
    ax.spines["bottom"].set_linewidth(0.6)


# ══════════════════════════════════════════════════════════════════
#  Step 5：绘图
# ══════════════════════════════════════════════════════════════════
_STAT_RECORDS: list[dict] = []   # 全局，统一收集后保存


def plot_group_bar_with_stats(stats: dict,
                               group_key: str,
                               model_name: str,
                               save_dir: str):
    stat_keys  = [f"{group_key}_{lb}" for lb in GROUP_LABELS[group_key]
                  if f"{group_key}_{lb}" in stats]
    sub_labels = [k.split("_", 1)[1] for k in stat_keys]
    n_grp      = len(stat_keys)
    x          = np.arange(n_grp)
    bar_colors = _GROUP_PALETTES[group_key][:n_grp]

    for mi, (metric, unit) in enumerate(zip(METRICS, UNITS)):
        fig, ax = plt.subplots(figsize=(6, 5), constrained_layout=True)


        means = [stats[k]["mae"][mi] for k in stat_keys]
        stds  = [stats[k]["std"][mi] for k in stat_keys]

        # ── 柱状图 ─────────────────────────────────────────────
        ax.bar(x, means,
               width=0.55,
               color=bar_colors,
               edgecolor="white",
               linewidth=2.5,
               alpha=0.92,
               yerr=stds,
               error_kw=dict(elinewidth=0.8, capsize=3,
                             capthick=0.8, ecolor="#444", alpha=0.6),
               zorder=3)

        # ── n= 标注 ──────────────────────────────────────────
        for xi, k in zip(x, stat_keys):
            n = stats[k]["n"]
            ax.text(xi, ax.get_ylim()[0] - (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.1,
                    f"n={n:,}",
                    ha="center", va="top",
                    # fontsize=5.5,
                    # color="#666666",
                    clip_on=False)
        # ── Kruskal-Wallis + Dunn/Mann-Whitney ───────────────
        groups_err = [stats[k]["err"][:, mi] for k in stat_keys]
        kw_p, dunn = run_kruskal_dunn(groups_err)
        stars_kw   = p_to_stars(kw_p)

        if dunn is not None:
            bar_tops = [m + s for m, s in zip(means, stds)]
            yrange   = ax.get_ylim()[1] - ax.get_ylim()[0]
            level_h  = yrange * 0.10
            occupied: dict[int, int] = {}

            pairs = list(itertools.combinations(range(n_grp), 2))
            # 显著对排前（占层级低），ns 排后
            pairs.sort(key=lambda ij: dunn.iloc[ij[0], ij[1]])

            for i, j in pairs:
                p_ij = dunn.iloc[i, j]
                s_ij = p_to_stars(p_ij)

                level = max(
                    (occupied.get(k, 0) for k in range(i, j + 1)),
                    default=0)
                y_ann = (max(bar_tops[i], bar_tops[j])
                         + level * level_h
                         + yrange * 0.09)

                _annotate_significance(ax, x[i], x[j], y_ann, s_ij)

                for k in range(i, j + 1):
                    occupied[k] = level + 1

                _STAT_RECORDS.append({
                    "model":           model_name,
                    "group_key":       group_key,
                    "metric":          metric,
                    "unit":            unit,
                    "subgroup_A":      sub_labels[i],
                    "subgroup_B":      sub_labels[j],
                    "n_A":             stats[stat_keys[i]]["n"],
                    "n_B":             stats[stat_keys[j]]["n"],
                    "MAE_A":           round(means[i], 3),
                    "MAE_B":           round(means[j], 3),
                    "KW_p":            float(kw_p),
                    "KW_stars":        stars_kw,
                    "KW_p_display":    format_p(kw_p),
                    "Dunn_p":          float(p_ij),
                    "Dunn_stars":      s_ij,
                    "Dunn_p_display":  format_p(p_ij),
                })

            if occupied:
                extra = max(occupied.values()) * level_h * 0.5
                ax.set_ylim(ax.get_ylim()[0],
                            ax.get_ylim()[1] + extra)

        # # KW p 标注（右上角，小字）
        # if not np.isnan(kw_p):
        #     ax.text(0.98, 0.99,
        #             f"KW p={format_p(kw_p)}",
        #             transform=ax.transAxes,
        #             ha="right", va="top",
        #             fontsize=5.5, color="#666", style="italic")

        _setup_ax(ax, f"BD {metric} ({unit})", sub_labels)

        for ext in ("pdf", "png"):
            fpath = os.path.join(save_dir,
                                 f"BD_{group_key}_{metric}.{ext}")
            fig.savefig(fpath, dpi=600, bbox_inches="tight")
        plt.close(fig)
        print(f"  📊 BD_{group_key}_{metric}.pdf")


# ══════════════════════════════════════════════════════════════════
#  Step 6：结果整理为 DataFrame → Excel
# ══════════════════════════════════════════════════════════════════
def stats_to_dataframe(stats: dict, model_name: str) -> pd.DataFrame:
    rows = []
    for key, s in stats.items():
        if key == "All":
            group_key = "All"
            subgroup  = "All"
        else:
            # key 格式固定为 "{group_key}_{label}"，label 本身可能含 "_"
            group_key, subgroup = key.split("_", 1)
        row = {
            "Model":    model_name,
            "Group":    group_key,
            "Subgroup": subgroup,
            "N":        s["n"],
        }
        for mi, (m, unit) in enumerate(zip(METRICS, UNITS)):
            row[f"{m}_MAE ({unit})"]     = round(s["mae"][mi], 3)
            row[f"{m}_STD"]              = round(s["std"][mi], 3)
            row[f"{m}_display"]          = (
                f"{s['mae'][mi]:.2f}±{s['std'][mi]:.2f}")
        rows.append(row)
    return pd.DataFrame(rows)


def _apply_excel_style(ws, df: pd.DataFrame):
    HDR_FILL  = PatternFill("solid", fgColor="1F497D")
    GRP_FILLS = {
        "age":      PatternFill("solid", fgColor="DEEAF1"),
        "bmi":      PatternFill("solid", fgColor="E2EFDA"),
        "gender":   PatternFill("solid", fgColor="FFF2CC"),
        "position": PatternFill("solid", fgColor="FCE4D6"),
        "All":      PatternFill("solid", fgColor="F2F2F2"),
    }
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=8)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border

    # 兼容两种 sheet：MAE_summary 用 "Group"，Statistics 用 "group_key"
    grp_col_name = "Group" if "Group" in df.columns else "group_key"
    group_col_idx = df.columns.get_loc(grp_col_name) + 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        gk   = ws.cell(row=row_idx, column=group_col_idx).value or "All"
        fill = GRP_FILLS.get(gk, GRP_FILLS["All"])
        for cell in row:
            cell.fill      = fill
            cell.border    = border
            cell.font      = Font(name="Arial", size=8)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, col in enumerate(df.columns, 1):
        w = 14 if "display" in col else 10
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════
#  主流程
# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # pkl_dir   =  your pkl path

    save_dir  = "../figures/subgroup_BD_subject_level/"
    save_xlsx = f"../results/subgroup_BD_subject_level_{formatted_date}.xlsx"
    os.makedirs(save_dir, exist_ok=True)

    all_result_dfs: dict[str, pd.DataFrame] = {}

    for model_name in MODELNAMES[3:4]:
        pkl_files = glob.glob(
            os.path.join(pkl_dir, model_name + "*_prediction.pkl"))
        if not pkl_files:
            print(f"[SKIP] {model_name}: no file found")
            continue

        raw_df = pd.read_pickle(pkl_files[0])
        print(f"\n▶ {model_name}  (n_segments={len(raw_df)})")

        # compute_subgroup_stats 内部自行处理 subject-level / segment-level 分组
        stats  = compute_subgroup_stats(raw_df)
        n_subj = stats["All"]["n"]
        print(f"   n_subjects={n_subj}"
              f"  (position 亚组人数之和可 > {n_subj}，属正常)")
        df     = stats_to_dataframe(stats, model_name)
        all_result_dfs[model_name] = df

        for gk in GROUPS.keys():
            plot_group_bar_with_stats(stats, gk, model_name, save_dir)

    # ── 保存 Excel ────────────────────────────────────────────────
    os.makedirs(os.path.dirname(save_xlsx), exist_ok=True)
    df_stat = pd.DataFrame(_STAT_RECORDS).sort_values(
        ["model", "group_key", "metric", "Dunn_p"]).reset_index(drop=True)
    df_all  = pd.concat(list(all_result_dfs.values()), ignore_index=True)

    with pd.ExcelWriter(save_xlsx, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="MAE_summary", index=False)
        df_stat.to_excel(writer, sheet_name="Statistics", index=False)

        # 样式
        wb = writer.book
        for sheet_name, df in [("MAE_summary", df_all),
                                ("Statistics",  df_stat)]:
            _apply_excel_style(wb[sheet_name], df)

    print(f"\n✅ 结果已保存至 {save_xlsx}")